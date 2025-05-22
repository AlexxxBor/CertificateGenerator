import os
import openpyxl

from enum import Enum
from docxtpl import DocxTemplate
from docx2pdf import convert
from art import tprint


class CertType(Enum):
    MAIN_CERT = "сертификат"
    DIST_CERT = "сертификат с отличием"


def get_template(cert_type: CertType) -> DocxTemplate:
    if cert_type == CertType.DIST_CERT:
        return DocxTemplate("templates/tpl_with_distinction.docx")
    return DocxTemplate("templates/tpl_certificate.docx")


def get_dir(directory: str):
    if not os.path.exists(directory):
        os.makedirs(directory)
        return directory
    return directory


def make_certificate(tpl_data: dict, cert_type: CertType, path: str) -> None:
    docx_dir = get_dir(f"{path}/docx")
    file_name = f"{tpl_data["surname"]} {tpl_data["name"]} {tpl_data["patronymic"]}"

    certificate = get_template(cert_type)
    certificate.render(tpl_data)
    certificate.save(f"{docx_dir}/{file_name}.docx")

    pdf_dir = get_dir(f"{path}/pdf")
    docx_file = f"{docx_dir}/{file_name}.docx"
    pdf_file = f"{pdf_dir}/{file_name}.pdf"
    convert(docx_file, pdf_file)


WORKING_DIR = get_dir("сертификаты")
CERT_DATA_SHEET = "cert_data"

tpl_data_keys = ("surname", "name", "patronymic", "course", "mod", "hour", "cert", "number")

wb = openpyxl.load_workbook(filename="data/IT-куб.xlsx")


def print_info(count: int):
    border = "█" * 50
    print(border)
    print("   Распечатано сертификатов всего:", count)
    print("   Из них:")
    for sheet in wb.sheetnames:
        if sheet == 'cert_data':
            continue

        active_sheet = wb[sheet]
        print(f"         ├ {sheet}: {len(list(active_sheet.iter_rows())) - 1}")
    print(border)


def main():
    tprint('starting...')
    count = 0

    for sheet in wb.sheetnames:
        if sheet == CERT_DATA_SHEET:
            continue

        try:
            course_dir = get_dir(f"{WORKING_DIR}/{sheet}")
        except OSError as e:
            print(f"Не могу создать папку курса в папке '{WORKING_DIR}' для листа '{sheet}'.")
            print(f"Возникла ошибка: {e}")
            break

        try:
            for row in wb[sheet].iter_rows(min_row=2):
                tpl_data_values = tuple(cell.value for cell in row)

                if None in tpl_data_values:
                    break

                context = {tpl_data_keys[i]: value for i, value in enumerate(tpl_data_values)}

                edu_module = context.pop('mod')
                if edu_module != "без модуля":
                    context['course'] = f"{context['course']} ({edu_module})"

                try:
                    make_certificate(context, CertType(context["cert"]), course_dir)
                    count += 1
                except Exception as e:
                    print(f"При создании сертификата возникла ошибка: {e}")

        except Exception as e:
            print(f"При формировании набора данных возникла ошибка: {e}")

    tprint('done!')
    print_info(count)
    input()

if __name__ == "__main__":
    main()
